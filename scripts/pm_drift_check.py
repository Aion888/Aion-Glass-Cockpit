from __future__ import annotations
from pathlib import Path
import re
import sys
import pandas as pd
from openpyxl import load_workbook

XLSX = Path("data/Project_Aion_PM_System.xlsx")
SPEC_DIR = Path("Project_Aion/01_Project_Framework/00_Master_Index/Framework_Directory_Spec")
TICKETS_SHEET = "04_Tickets"
REQUIRED = ["Realm", "Framework_Path", "Roadmap_Milestone"]

BLANK_STRINGS = {"", "nan", "none", "null"}

def allowed_realms() -> set[str]:
    return {p.stem for p in SPEC_DIR.glob("*.md")}

def guess_header_row(ws, scan_rows: int = 30, max_cols: int = 80) -> int:
    best_row = 1
    best_score = (-1, -1)
    for r in range(1, min(scan_rows, ws.max_row) + 1):
        vals = []
        for c in range(1, max_cols + 1):
            v = ws.cell(row=r, column=c).value
            vals.append("" if v is None else str(v).strip())

        while vals and vals[-1] == "":
            vals.pop()

        nonempty = sum(v != "" for v in vals)
        if nonempty == 0:
            continue

        stringy = sum(v != "" and not re.fullmatch(r"[-+]?\d+(\.\d+)?", v) for v in vals)
        score = (nonempty, stringy)
        if score > best_score:
            best_score = score
            best_row = r
    return best_row

def add_required_columns(ws, header_row: int) -> list[str]:
    headers = []
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=header_row, column=c).value
        headers.append("" if v is None else str(v).strip())

    last = 0
    for i, h in enumerate(headers, start=1):
        if h != "":
            last = i

    existing = set(h for h in headers if h != "")
    added = []
    for col in REQUIRED:
        if col not in existing:
            last += 1
            ws.cell(row=header_row, column=last).value = col
            added.append(col)
    return added

def norm_cell(v) -> str:
    if pd.isna(v):
        return ""
    s = str(v).strip()
    return "" if s.lower() in BLANK_STRINGS else s

def is_blank(v) -> bool:
    return norm_cell(v) == ""

def main():
    if not XLSX.exists():
        print(f"ERROR: missing {XLSX}")
        sys.exit(1)

    wb = load_workbook(XLSX)
    if TICKETS_SHEET not in wb.sheetnames:
        print("ERROR: expected sheet not found:", TICKETS_SHEET)
        print("Sheets:", wb.sheetnames)
        sys.exit(1)

    ws = wb[TICKETS_SHEET]
    header_row = guess_header_row(ws)
    print("Sheets:", wb.sheetnames)
    print("Tickets sheet:", TICKETS_SHEET)
    print("Detected header row:", header_row)

    added = add_required_columns(ws, header_row)
    if added:
        wb.save(XLSX)
        print("Added missing alignment columns:", added)
    wb.close()

    df = pd.read_excel(XLSX, sheet_name=TICKETS_SHEET, engine="openpyxl", header=header_row - 1)
    df.columns = [str(c).strip() for c in df.columns]

    missing_cols = [c for c in REQUIRED if c not in df.columns]
    if missing_cols:
        print("\nStill missing required columns:", missing_cols)
        print("Columns found:", list(df.columns))
        sys.exit(2)

    realms = allowed_realms()

    content_cols = [c for c in df.columns if c not in REQUIRED]

    report_rows = []
    for idx, row in df.iterrows():
        # Skip fully empty rows (ignoring the 3 alignment columns)
        if all(is_blank(row.get(c)) for c in content_cols):
            continue

        realm_raw = row.get("Realm")
        fpath_raw = row.get("Framework_Path")
        mile_raw  = row.get("Roadmap_Milestone")

        realm = norm_cell(realm_raw)
        fpath = norm_cell(fpath_raw)
        mile  = norm_cell(mile_raw)

        issues = []
        if realm == "": issues.append("Realm empty")
        if fpath == "": issues.append("Framework_Path empty")
        if mile  == "": issues.append("Roadmap_Milestone empty")

        if realm and realm not in realms:
            issues.append(f"Realm not in spec ({realm})")

        if fpath and not Path(fpath).exists():
            issues.append("Framework_Path does not exist on disk")

        if issues:
            report_rows.append({
                "Row": idx + 1 + header_row,  # approx Excel row number
                "Realm": realm,
                "Framework_Path": fpath,
                "Roadmap_Milestone": mile,
                "Issues": "; ".join(issues),
            })

    out = Path("data/pm_drift_report.csv")
    pd.DataFrame(report_rows).to_csv(out, index=False)

    print(f"\nDrift report written: {out}")
    print(f"Issues found: {len(report_rows)}")
    if len(report_rows) == 0:
        print("✅ No drift detected for required alignment fields.")

if __name__ == "__main__":
    main()
