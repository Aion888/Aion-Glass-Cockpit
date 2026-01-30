import sys, csv
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation

def find_sheet(wb, keywords):
    kws = [k.lower() for k in keywords]
    for name in wb.sheetnames:
        n = name.lower()
        if all(k in n for k in kws):
            return name
    return None

def ensure_col(ws, header, header_row=1):
    headers = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=header_row, column=c).value
        if isinstance(v, str) and v.strip():
            headers[v.strip().lower()] = c

    key = header.lower()
    if key in headers:
        return headers[key]

    new_col = ws.max_column + 1
    ws.cell(row=header_row, column=new_col).value = header
    return new_col

def read_framework_nodes(csv_path: Path):
    with csv_path.open("r", encoding="utf-8", newline="") as f:
        r = csv.DictReader(f)
        if not r.fieldnames or "Framework_Node" not in r.fieldnames:
            raise SystemExit("Framework_Index.csv missing column: Framework_Node")
        nodes = []
        for row in r:
            n = (row.get("Framework_Node") or "").strip()
            if n:
                nodes.append(n)
    # de-dupe preserving order
    seen = set()
    nodes = [n for n in nodes if not (n in seen or seen.add(n))]
    return nodes

def main(xlsx_path):
    xlsx_path = Path(xlsx_path).expanduser().resolve()
    if not xlsx_path.exists():
        raise SystemExit(f"Workbook not found: {xlsx_path}")

    csv_path = Path("Framework_Index.csv")
    if not csv_path.exists():
        raise SystemExit("Framework_Index.csv not found. Generate it first.")

    nodes = read_framework_nodes(csv_path)
    if not nodes:
        raise SystemExit("No Framework nodes found in Framework_Index.csv")

    wb = load_workbook(xlsx_path)

    # 1) write Framework_Index sheet
    if "Framework_Index" in wb.sheetnames:
        wb.remove(wb["Framework_Index"])
    ws_idx = wb.create_sheet("Framework_Index")
    ws_idx["A1"] = "Framework_Node"
    for i, n in enumerate(nodes, start=2):
        ws_idx[f"A{i}"] = n

    # 2) find likely EPICS + Tickets sheets
    epics_sheet   = find_sheet(wb, ["epic"]) or find_sheet(wb, ["roadmap"]) or find_sheet(wb, ["epics"])
    tickets_sheet = find_sheet(wb, ["ticket"]) or find_sheet(wb, ["tickets"]) or find_sheet(wb, ["backlog"])

    targets = []
    if epics_sheet:
        targets.append(epics_sheet)
    if tickets_sheet and tickets_sheet != epics_sheet:
        targets.append(tickets_sheet)

    if not targets:
        print("Sheetnames found:", wb.sheetnames)
        raise SystemExit("Couldn't auto-find EPICS/Tickets sheets. Rename them to include 'Epic'/'Roadmap' and 'Ticket'.")

    # 3) dropdown validation: Framework_Index!$A$2:$A$N
    max_row = len(nodes) + 1
    dv = DataValidation(
        type="list",
        formula1=f"=Framework_Index!$A$2:$A${max_row}",
        allow_blank=True
    )

    for sheet_name in targets:
        ws = wb[sheet_name]
        col = ensure_col(ws, "Framework_Node", header_row=1)
        ws.add_data_validation(dv)
        dv.add(f"{ws.cell(row=2, column=col).coordinate}:{ws.cell(row=5000, column=col).coordinate}")

    out = xlsx_path.with_name(xlsx_path.stem + "_ALIGNED.xlsx")
    wb.save(out)
    print("Wrote:", out)
    print("Updated sheets:", targets)

if __name__ == "__main__":
    if len(sys.argv) < 2:
        raise SystemExit("Usage: python tools/sync_framework_to_pm.py path/to/PM.xlsx")
    main(sys.argv[1])
