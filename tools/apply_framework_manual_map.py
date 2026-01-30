import csv, sys
from pathlib import Path
from openpyxl import load_workbook

def norm(s): return (s or "").strip().lower()

def headers_map(ws):
    m={}
    for c in range(1, ws.max_column+1):
        v = ws.cell(1,c).value
        if isinstance(v,str) and v.strip():
            m[norm(v)] = c
    return m

def ensure_col(ws, header):
    h = headers_map(ws)
    key = norm(header)
    if key in h:
        return h[key]
    c = ws.max_column + 1
    ws.cell(1,c).value = header
    return c

def main(xlsx_in, csv_map, xlsx_out):
    xlsx_in = Path(xlsx_in).resolve()
    csv_map = Path(csv_map).resolve()
    xlsx_out = Path(xlsx_out).resolve()

    wb = load_workbook(xlsx_in)
    updated = 0

    with csv_map.open("r", encoding="utf-8", newline="") as f:
        r = csv.DictReader(f)
        for row in r:
            sheet = (row.get("sheet") or "").strip()
            rr = row.get("row")
            final = (row.get("framework_node_final") or "").strip()
            if not sheet or not rr or not final:
                continue
            if sheet not in wb.sheetnames:
                continue
            try:
                rr = int(rr)
            except Exception:
                continue
            ws = wb[sheet]
            c_node = ensure_col(ws, "Framework_Node")
            ws.cell(rr, c_node).value = final
            updated += 1

    wb.save(xlsx_out)
    print("Input :", xlsx_in)
    print("Map   :", csv_map)
    print("Output:", xlsx_out)
    print("Rows updated:", updated)

if __name__ == "__main__":
    if len(sys.argv) < 4:
        raise SystemExit("Usage: python tools/apply_framework_manual_map.py <in.xlsx> <map.csv> <out.xlsx>")
    main(sys.argv[1], sys.argv[2], sys.argv[3])
