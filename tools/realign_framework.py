import re
import sys
import argparse
from pathlib import Path
from difflib import SequenceMatcher
from collections import defaultdict

from openpyxl import load_workbook

# ----------------------------
# Helpers
# ----------------------------
def norm(s: str) -> str:
    return re.sub(r"\s+", " ", (s or "").strip().lower())

def tokenize_node(node: str):
    # "02_Technology_Advantage" -> ["technology","advantage"]
    node = re.sub(r"^\d{2}_", "", node.strip())
    parts = re.split(r"[_\-/\s]+", node)
    toks = []
    for p in parts:
        p = re.sub(r"[^a-zA-Z0-9]", "", p).lower()
        if len(p) >= 3 and not p.isdigit():
            toks.append(p)
    # de-dupe preserve order
    seen = set()
    return [t for t in toks if not (t in seen or seen.add(t))]

def best_text_columns(headers):
    # heuristic keywords
    kw = [
        "epic", "goal", "title", "summary", "description", "details", "notes",
        "deliverable", "outcome", "scope", "objective", "problem", "solution",
        "task", "ticket", "feature", "story"
    ]
    cols = []
    for h, col in headers.items():
        if any(k in h for k in kw) and "framework_" not in h:
            cols.append(col)
    return sorted(set(cols))

def headers_map(ws, header_row=1):
    m = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=header_row, column=c).value
        if isinstance(v, str) and v.strip():
            m[norm(v)] = c
    return m

def ensure_col(ws, header, header_row=1):
    hmap = headers_map(ws, header_row=header_row)
    key = norm(header)
    if key in hmap:
        return hmap[key]
    new_col = ws.max_column + 1
    ws.cell(row=header_row, column=new_col).value = header
    return new_col

def build_row_text(ws, r, cols, max_chars=1200):
    parts = []
    for c in cols:
        v = ws.cell(row=r, column=c).value
        if isinstance(v, str) and v.strip():
            parts.append(v.strip())
    text = " | ".join(parts)
    text = re.sub(r"\s+", " ", text).strip().lower()
    return text[:max_chars]

def score_node(node, tokens, text):
    if not text:
        return 0.0
    total = sum(len(t) for t in tokens) or 1
    hit = sum(len(t) for t in tokens if t in text)
    base = hit / total  # 0..1
    node_words = " ".join(tokens) if tokens else node.lower()
    fuzzy = SequenceMatcher(None, node_words, text[:500]).ratio()  # 0..1
    return 0.7 * base + 0.3 * fuzzy

def newest_xlsx(root: Path):
    files = list(root.rglob("*.xlsx"))
    files = [p for p in files if not p.name.endswith("_MAPPED.xlsx")]
    if not files:
        raise SystemExit("No .xlsx found under this folder.")
    return max(files, key=lambda p: p.stat().st_mtime)

# ----------------------------
# Main
# ----------------------------
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("xlsx", nargs="?", help="Input workbook (defaults to newest .xlsx found)")
    ap.add_argument("--out", default=None, help="Output workbook path")
    ap.add_argument("--min_conf", type=float, default=0.55, help="Minimum confidence to auto-fill Framework_Node")
    ap.add_argument("--overwrite", action="store_true", help="Overwrite existing Framework_Node values")
    args = ap.parse_args()

    root = Path(".").resolve()
    xlsx_path = Path(args.xlsx).expanduser().resolve() if args.xlsx else newest_xlsx(root)
    if not xlsx_path.exists():
        raise SystemExit(f"Workbook not found: {xlsx_path}")

    wb = load_workbook(xlsx_path)

    # Framework nodes
    if "Framework_Index" not in wb.sheetnames:
        raise SystemExit("Missing sheet 'Framework_Index'. Run your earlier alignment step first.")
    ws_idx = wb["Framework_Index"]

    nodes = []
    for r in range(2, ws_idx.max_row + 1):
        v = ws_idx.cell(row=r, column=1).value
        if isinstance(v, str) and v.strip():
            nodes.append(v.strip())
    if not nodes:
        raise SystemExit("Framework_Index has no nodes.")

    node_tokens = {n: tokenize_node(n) for n in nodes}

    # Target sheets (your known names)
    targets = [s for s in ["02_Roadmap", "04_Tickets"] if s in wb.sheetnames]
    if not targets:
        raise SystemExit(f"Couldn't find target sheets. Found: {wb.sheetnames}")

    # Coverage tracking
    epic_counts = defaultdict(int)
    ticket_counts = defaultdict(int)
    epic_ids_by_node = defaultdict(list)
    ticket_ids_by_node = defaultdict(list)

    for sheet_name in targets:
        ws = wb[sheet_name]
        hmap = headers_map(ws)

        # Identify relevant text columns
        text_cols = best_text_columns(hmap)
        if not text_cols:
            # fallback: use all string-ish columns except framework columns
            for h, c in hmap.items():
                if "framework_" not in h:
                    text_cols.append(c)
            text_cols = sorted(set(text_cols))

        # Core columns
        c_node = ensure_col(ws, "Framework_Node", 1)
        c_sug  = ensure_col(ws, "Framework_Suggested", 1)
        c_conf = ensure_col(ws, "Framework_Confidence", 1)
        c_top3 = ensure_col(ws, "Framework_Top3", 1)

        # Try to find ID-ish columns for reporting
        c_id = None
        for key in ["epic id", "ticket id", "id"]:
            if key in hmap:
                c_id = hmap[key]
                break

        # Iterate rows
        for r in range(2, ws.max_row + 1):
            existing = ws.cell(row=r, column=c_node).value
            existing = existing.strip() if isinstance(existing, str) else ""

            text = build_row_text(ws, r, text_cols)
            if not text:
                # still write blanks for transparency
                ws.cell(row=r, column=c_sug).value = ""
                ws.cell(row=r, column=c_conf).value = 0.0
                ws.cell(row=r, column=c_top3).value = ""
                continue

            scored = []
            for n in nodes:
                sc = score_node(n, node_tokens[n], text)
                if sc > 0:
                    scored.append((sc, n))
            scored.sort(reverse=True)

            top = scored[:3]
            sug = top[0][1] if top else ""
            conf = float(top[0][0]) if top else 0.0
            top3 = ", ".join([f"{n} ({sc:.2f})" for sc, n in top])

            ws.cell(row=r, column=c_sug).value = sug
            ws.cell(row=r, column=c_conf).value = round(conf, 3)
            ws.cell(row=r, column=c_top3).value = top3

            should_fill = (args.overwrite or not existing) and (conf >= args.min_conf) and bool(sug)
            if should_fill:
                ws.cell(row=r, column=c_node).value = sug
                existing = sug  # for coverage

            # Coverage bookkeeping (only if assigned)
            assigned = existing
            if assigned:
                row_id = ""
                if c_id:
                    v = ws.cell(row=r, column=c_id).value
                    row_id = str(v).strip() if v is not None else ""
                if sheet_name == "02_Roadmap":
                    epic_counts[assigned] += 1
                    if row_id:
                        epic_ids_by_node[assigned].append(row_id)
                else:
                    ticket_counts[assigned] += 1
                    if row_id:
                        ticket_ids_by_node[assigned].append(row_id)

    # Build coverage sheet
    if "Framework_Coverage" in wb.sheetnames:
        wb.remove(wb["Framework_Coverage"])
    ws_cov = wb.create_sheet("Framework_Coverage")
    ws_cov.append(["Framework_Node", "Epic_Count", "Ticket_Count", "Epic_IDs", "Ticket_IDs"])

    for n in nodes:
        e = epic_counts.get(n, 0)
        t = ticket_counts.get(n, 0)
        eids = ", ".join(epic_ids_by_node.get(n, [])[:50])
        tids = ", ".join(ticket_ids_by_node.get(n, [])[:50])
        ws_cov.append([n, e, t, eids, tids])

    # Build “needs review” sheets
    def make_review_sheet(name, src_sheet):
        if name in wb.sheetnames:
            wb.remove(wb[name])
        ws_out = wb.create_sheet(name)
        ws_in = wb[src_sheet]
        h = headers_map(ws_in)
        # grab columns if present
        c_node = h.get("framework_node")
        c_sug  = h.get("framework_suggested")
        c_conf = h.get("framework_confidence")
        c_top3 = h.get("framework_top3")
        ws_out.append(["Row", "Framework_Node", "Suggested", "Confidence", "Top3"])
        for r in range(2, ws_in.max_row + 1):
            node = ws_in.cell(row=r, column=c_node).value if c_node else ""
            sug  = ws_in.cell(row=r, column=c_sug).value if c_sug else ""
            conf = ws_in.cell(row=r, column=c_conf).value if c_conf else 0.0
            try:
                conf_f = float(conf or 0.0)
            except Exception:
                conf_f = 0.0
            top3 = ws_in.cell(row=r, column=c_top3).value if c_top3 else ""

            node_s = node.strip() if isinstance(node, str) else ""
            sug_s  = sug.strip() if isinstance(sug, str) else ""
            if (not node_s) or (conf_f < args.min_conf):
                ws_out.append([r, node_s, sug_s, conf_f, top3])

    if "02_Roadmap" in wb.sheetnames:
        make_review_sheet("Review_Roadmap", "02_Roadmap")
    if "04_Tickets" in wb.sheetnames:
        make_review_sheet("Review_Tickets", "04_Tickets")

    # Save output
    out = Path(args.out).expanduser().resolve() if args.out else xlsx_path.with_name(xlsx_path.stem + "_MAPPED.xlsx")
    wb.save(out)
    print("Input :", xlsx_path)
    print("Output:", out)
    print("Sheets updated:", targets)
    print("Review sheets: Review_Roadmap, Review_Tickets")
    print("Coverage sheet: Framework_Coverage")

if __name__ == "__main__":
    main()
